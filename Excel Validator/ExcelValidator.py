import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


class ValidateQuestionnaire:
    def __init__(self, Template=None, Filled_Questionnaire=None):
        if Template is None:
            raise ValueError("Questionnaire Template cannot be Null.")
        if Filled_Questionnaire is None:
            raise ValueError("Filled Questionnaire cannot be Null.")

        
        Template_WB = Workbook()
        Template_WS = Template_WB.active
        Template_WB.title = "Sheet 1"
        for row in dataframe_to_rows(Template.where(Template.notnull(), None), index=False, header=True):
            Template_WS.append(row)

        Filled_Questionnaire_WB = Workbook()
        Filled_Questionnaire_WS = Filled_Questionnaire_WB.active
        Filled_Questionnaire_WB.title = "Sheet 1"
        for row in dataframe_to_rows(Filled_Questionnaire.where(Filled_Questionnaire.notnull(), None), index=False, header=True):
            Filled_Questionnaire_WS.append(row)


        self.template = Template_WB
        self.filled_questionnaire = Filled_Questionnaire_WB
        self.template_sheet = Template_WS
        self.filled_questionnaire_sheet = Filled_Questionnaire_WS

        self.not_applicable = ("", "na", "n/a", "not applicable")

        self.template_headers = None
        self.filled_questionnaire_headers = None
        self.detect_columns()

        self.template_questions_column = self.template_headers.get("question", self.template_headers.get("questions"))
        self.template_validation_colum = self.template_headers.get("validation", self.template_headers.get("validations"))
        self.filled_questionnaire_questions_column = self.filled_questionnaire_headers.get("question", self.filled_questionnaire_headers.get("questions"))
        self.filled_questionnaire_answers_column = self.filled_questionnaire_headers.get("answer", self.filled_questionnaire_headers.get("answers"))

        self.validate_questions()
        self.validation_errors = []
        self.pick_groups = {}
        self.validate()

    def detect_columns(self):
        template_headers, filled_questionnaire_headers = {}, {}
        for col in range(1, self.template_sheet.max_column + 1):
            val = self.template_sheet.cell(row=1, column=col).value
            if val:
                template_headers[val.strip().lower()] = col
        for col in range(1, self.filled_questionnaire_sheet.max_column + 1):
            val = self.filled_questionnaire_sheet.cell(row=1, column=col).value
            if val:
                filled_questionnaire_headers[val.strip().lower()] = col
        self.template_headers = template_headers
        self.filled_questionnaire_headers = filled_questionnaire_headers
        return

    def validate_questions(self):
        template_questions = [
            self.template_sheet.cell(row=row, column=self.template_questions_column).value
            for row in range(2, self.template_sheet.max_row + 1)
        ]
        filled_questionnaire_question = [
            self.filled_questionnaire_sheet.cell(row=row, column=self.filled_questionnaire_questions_column).value
            for row in range(2, self.filled_questionnaire_sheet.max_row + 1)
        ]

        if template_questions != filled_questionnaire_question:
            raise ValueError("Submitted questionnaire does not match template. Kindly get updated template from cloud "
                             "team.")
        return

    def parse_validators(self, cell_value):
        if cell_value is None:
            return []

        raw = str(cell_value)
        parts = [p.strip().lower() for p in raw.split(";")]
        parsed = []

        for p in parts:
            if p.startswith("list"):
                items = p[p.index("[") + 1:p.rindex("]")].split(",")
                parsed.append(("LIST", [i.strip() for i in items]))

            elif p.startswith("regex"):
                pat = p[p.index("[") + 1:p.rindex("]")].strip('"')
                parsed.append(("REGEX", pat))

            elif p.startswith("pick"):
                nums = p[p.index("[") + 1:p.rindex("]")].split(",")
                parsed.append(("PICK", [int(n.strip()) for n in nums]))

            else:
                parsed.append((p.upper(), None))

        return parsed

    def validate(self):
        for row in range(2, self.template_sheet.max_row + 1):  # Iterating over every row.
            value = self.filled_questionnaire_sheet.cell(row=row, column=self.filled_questionnaire_answers_column).value
            validation = self.parse_validators(self.template_sheet.cell(row=row, column=self.template_validation_colum).value)

            if value is None and "NULL" not in validation:
                msg = "Value cannot be left empty. If this answer is not required, kindly put 'N/A' instead."
                self.validation_errors.append(f"Row {row}: {msg}")
                continue

            for vtype, arg in validation:

                if vtype == "TEXT":
                    ok, msg = self.is_text(value)

                elif vtype == "NUMBER":
                    ok, msg = self.is_number(value)

                elif vtype == "NEGATIVENUMBER":
                    ok, msg = self.is_negetive_number(value)

                elif vtype == "YES/NO":
                    ok, msg = self.is_yes_no(value)

                elif vtype == "LIST":
                    ok, msg = self.is_list(value, items=arg)

                elif vtype == "REGEX":
                    ok, msg = self.is_regex(value, regex=arg)

                elif vtype == "PICK":
                    self.pick_groups.setdefault(tuple(arg), []).append((row, value))
                    ok, msg = True, None

                elif vtype == "NULL":
                    ok, msg = self.is_null(value)

                else:
                    ok, msg = False, f"Unknown validation: {vtype}"

                if not ok:
                    self.validation_errors.append(f"Row {row}: {msg}")
                    break

        self.is_pick()


        if self.validation_errors:
            print("❌ Validation Failed:\n" + "\n".join(self.validation_errors))
            #raise ValueError("❌ Validation Failed:\n" + "\n".join(self.validation_errors))
        else:
            print("✅ Validation Successful.")

    def is_text(self, value):
        if not re.match(r"^(?=.*[a-zA-Z])[\x20-\x7E\s]*$", str(value)):
            return False, "Text conditions not satisfied. Text must contain atleast 1 letter. " \
            "Text cannot contain any non keyboard special characters." \
            " If not required, write N/A."
        return True, None

    def is_number(self, value):
        try:
            if str(value).strip().lower() in self.not_applicable:
                return True, None
            elif float(value) >= 0:
                return True, None
            else:
                return False, "Value must be greater then 0."
        except:
            return False, "Value must be a number. If not required, write N/A."

    def is_negetive_number(self, value):
        try:
            if str(value).strip().lower() in self.not_applicable:
                return True, None
            val = float(value)
            return True, None
        except:
            return False, "Value must be a number. If not required, write N/A."

    def is_yes_no(self, value):
        if not isinstance(value, str):
            return False, "Must be Yes or No"
        if str(value).strip().lower() in self.not_applicable:
            return True, None
        elif str(value).strip().lower() in ("yes", "no", "y", "n"):
            return True, None
        return False, "Value must be Yes or No. If not required, write N/A."

    def is_list(self, value, items):
        if not isinstance(value, str):
            return False, f"Must be one of {items}"
        if value.lower().strip() in [x.lower() for x in items]:
            return True, None
        return False, f"Value must be one of {items}. N/A is not an option for this question."

    def is_regex(self, value, regex):
        if bool(re.match(regex, str(value))):
            return True, None
        return False, "Regex match failed. Contact cloud team for help."

    def is_pick(self):
        for k, values in enumerate(self.pick_groups.keys()):
            groups = self.pick_groups[values]
            if len([v[1] for v in groups if str(v[1]).strip().lower() not in self.not_applicable]) != 1:
                self.validation_errors.append(f"PICK one from {values}: exactly ONE of the questions on these rows must be answered. Rest should be 'N/A'")

    def is_null(self, value):
        if value is None or str(value).strip().lower() in self.not_applicable:
            return True, None
        else:
            return False, "This answer must be left empty."


def main():
    TEMPLATE = "D:\\GitHub\\ExcelValidator\\ExcelValidator\\Excel Validator\\Questionnaire_Template.xlsx"  # Variables for testing.
    SUBMISSION = "D:\\GitHub\\ExcelValidator\\ExcelValidator\\Excel Validator\\Submission1.xlsx"

    #template_workbook = load_workbook(TEMPLATE)  # Loading of workbook needs to be done before class object creation.
    #submission_workbook = load_workbook(SUBMISSION)
    template_workbook = pd.read_excel(TEMPLATE, sheet_name=0, engine="openpyxl", keep_default_na=False)
    submission_workbook = pd.read_excel(SUBMISSION, sheet_name=0, engine="openpyxl", keep_default_na=False)

    validator = ValidateQuestionnaire(template_workbook, submission_workbook)


if __name__ == "__main__":
    main()
